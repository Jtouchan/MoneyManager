# the functions in this class carry out data analysis regarding spendings and savings.
import openpyxl as xl
import string
from time import strptime
import math
letters = list(string.ascii_uppercase)

class PreAnalysis :

    def Fixed_Values (ID_values_summed,Write_workbook, month):

        month_num = strptime(month,'%b').tm_mon
        if (month_num > 5):
            Rent = ##
        else:
            Rent = ##

        # sum all expenses from all catagories:
        Expenses = sum(ID_values_summed.values())
        Col = 'C'
        Row = '19'
        SheetName = "Fixed Numbers"
        Write_sheet=Write_workbook.get_sheet_by_name(SheetName)
        Net_Income = (Write_sheet[Col+Row].value)
        # How much money have you saved
        Savings = Net_Income - (Expenses + Rent)

        return Savings , Rent , Net_Income


    def Previous_Expenses (Write_workbook):
        Col = letters[2:11]
        Row = '6'
        Monthly_Expenses =[]
        temp=[]
        months_names = []
        for i in range(2,len(Write_workbook.sheetnames)):

            months_names.append(Write_workbook.sheetnames[i])
            Write_sheet=Write_workbook.get_sheet_by_name(Write_workbook.sheetnames[i])

            for ii in range (len(Col)):
                temp.append( Write_sheet[Col[ii]+Row].value)

            Monthly_Expenses.append(sum(temp))
            temp=[]

        return Monthly_Expenses,months_names


    def RemoveZero (ID_values_summed):
        Catagories_Zero = list(ID_values_summed.keys())
        Expenses_Zero = list(ID_values_summed.values())
        Expenses=list()
        Catagories = list()
        for i in range(len(Expenses_Zero)):
            if Expenses_Zero[i]==0:
                continue
            Expenses.append(Expenses_Zero[i])
            Catagories.append(Catagories_Zero[i])

        return Catagories , Expenses
