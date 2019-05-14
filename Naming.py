import string
import os
import numpy as np
import openpyxl as xl
import re
import calendar
import pyexcel as p



class PreProcess:
    def Naming(self):
        Mainpth = "C:\\Users\\gad-t\\Desktop\\Files\\Money Manger\\Bank Statements\\"
        WriteFileName= "Processed\\Combined\\Spending Summary.xlsx"
        Write_Full_pth =Mainpth+WriteFileName
        Write_workbook = xl.load_workbook(Write_Full_pth)
        fileExist=True

        letters = list(string.ascii_lowercase)
        Number_banks=2;

        for i in range(Number_banks):

            if (i==0):
                Bank = "UNFCU"
                FileName = "download.xlsx"
                SheetName="Download"
                Date_Label = "Transaction Date"

            if (i==1):
                Bank = "AMEX"
                FileName = "Summary.xls"
                SheetName="Transaction Details"
                Date_Label = "Date"

            Folderpth = Mainpth + "Pre-Proccessed\\"+ Bank +"\\"
            Filepth = Folderpth + FileName

            if ( os.path.isfile(Filepth)==0):  # if file does not exist skip it.
                fileExist=False
                continue

            ## Transfrom from xls to xlsx
            if (Bank == "AMEX"):
                p.save_book_as( file_name= Folderpth + 'Summary.xls',  dest_file_name= Folderpth+"Summary.xlsx")
                os.remove(Folderpth + 'Summary.xls')
                FileName = "Summary.xlsx"
                Filepth = Folderpth + FileName
                Bank_statement = xl.load_workbook(Filepth)
                ## remove extra sheet
                removed_sheet=Bank_statement.get_sheet_by_name('Transaction Summary')
                Bank_statement.remove_sheet(removed_sheet)
                Bank_statement.save (Filepth)
            Bank_statement = xl.load_workbook(Filepth)
            Read_sheet=Bank_statement.get_sheet_by_name(SheetName)

            ## find where the header row is
            Header_Finder  = Read_sheet['A']
            for labels_row in range(len(Header_Finder)):
                if Header_Finder[labels_row].value is not None:
                    break
            Read_Header =Read_sheet[labels_row+1]

            ## find the column containing the date info
            for i in range(len(Read_Header)):
                if (Read_Header[i].value == Date_Label):
                        break


            Transaction_Date =Read_sheet[letters[i]]
            month =  re.split("\/",Transaction_Date[int (labels_row+1)].value)[0]

            month = calendar.month_abbr[int(month)]

            New_Name = month + ".xlsx"

            Read_sheet.title = month
            Savepth = Mainpth + "Processed\\Individual\\"+ Bank + "\\"
            Bank_statement.save(Savepth + New_Name)

            # Delete Read File
            if (os.path.isfile(Filepth)):
                os.remove(Filepth)

        if ( fileExist==False) :
            month = Write_workbook.sheetnames[-1] # if file does not exist, look at the last sheet.

        return (month)
