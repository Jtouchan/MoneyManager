'''
CODE AUTHOR: GAD TOUCHAN
'''

###################
## Importing relevant libs
###################
import string
import os
import numpy as np
import openpyxl as xl
import re
import calendar
from BackEnd import BackEnd
from Naming import PreProcess
from Analysis import *
from tkinter import *
import tkinter as ttk

###################
##  Setting global variables and pre-processing statements
###################

PreProcess = PreProcess( ) ## Creating month-based named files.
BackEnd=BackEnd() ##  Reading raw data and GUI
Analysis = Analysis ( ) ## Analysis data and creating graphs



letters = list(string.ascii_lowercase)
month = PreProcess.Naming ( ) ## pre-process files and gets the month
#month = "Mar" # NOTE: if you want to analyse a specific month, you can overide here.

#skipable = [" Interest Charge on Cash Advances" , " Interest Charge on Purchases" , " AUTOMATIC PAYMENT" , " ONLINE PAYMENT" , " AUTOPAY PAYMENT"] ## move to excel

Mainpth = 'C:\\Users\\gad-t\\Desktop\\Files\\Money Manger\\Bank Statements\\Processed\\'
WriteFileName= "Combined\\Spending Summary.xlsx"
Write_Full_pth =Mainpth+WriteFileName

###################
## Setting-up Spending Summary file and getting expenses catagories
###################

# create the file if does not exist
if (not os.path.isfile(Write_Full_pth)):
    Write_workbook = xl.Workbook()
    Write_workbook.save(Write_Full_pth)

Write_workbook = xl.load_workbook(Write_Full_pth)

SheetName= "Catagories"
Skip_Sheet=Write_workbook.get_sheet_by_name(SheetName)
Skip_Column = Skip_Sheet['L']
skipable = [ ]
for i in range(1,len(Skip_Column)-1):
	if (Skip_Column[i].value != None):
		skipable.append(Skip_Column[i].value)

 
# create a sheet for the month and set this sheet as active
if (month not in Write_workbook.sheetnames):
    Write_workbook.create_sheet(month)
    Write_workbook.save(Write_Full_pth)

# read the expenses catagories.

ID,_ = BackEnd.DictCreater(Write_Full_pth)
ID_values= {k: [] for k in ID} # create another dictionary to contain numbers only
ID_values_summed = ID_values  #maybe delete this once you figured out the bug

for i in range(2):

    if (i==0):
        Bank= "UNFCU"
        Date_Label = "Transaction Date"
        Description_Label = "Transaction Description"
        Amount_Label = "Amount"
        id_1 = 1
        id_2 = 2

    if (i==1):
        Bank ="AMEX"
        Date_Label = "Date"
        Description_Label = "Description"
        Amount_Label = "Amount"
        id_1 = 0
        id_2 = 1

    ###################
    ## Paths
    ###################

    ReadFileName= "individual\\" + Bank + "\\" +month + ".xlsx"
    Read_Full_pth= Mainpth + ReadFileName

    ###################
    ## Setting-up Reading File
    ###################
    Read_workbook = xl.load_workbook(Read_Full_pth)
    SheetName= month
    Read_sheet=Read_workbook.get_sheet_by_name(SheetName)

    Header_Finder  = Read_sheet['A']
    for labels_row in range(len(Header_Finder)):
        if Header_Finder[labels_row].value is not None:
            break
    Read_Header =Read_sheet[labels_row+1]


    ###################
    ## Extract Transaction Description
    ###################
    for x in range(len(Read_Header)):
        if (Read_Header[x].value == Description_Label):
            break

    Transaction_Description =Read_sheet[letters[x]]

    ###################
    ## Extract Transaction Amount
    ###################
    for x in range(len(Read_Header)):
        if (Read_Header[x].value == Amount_Label):
            break

    Transaction_Amount = Read_sheet[letters[x]]

    ###################
    ## Extract Identifier
    ###################

    for x in range(len(Transaction_Description)-1):
        if (x >=labels_row):
            identifier = BackEnd.Identifier_cleaner(Transaction_Description,x,id_1,id_2)
            for type, subtype in ID.items(): # check if identifier exists

                NewValue = not any(identifier in val for val in ID.values()) # is this already exists ?
                skip = identifier  in skipable # should it be skipped ?

                if ( not NewValue and not skip ):
                    if (identifier in ID[type]):
                        ID_values[type].append(Transaction_Amount[x+1].value)  # get the x value of it.
                if ( NewValue and not skip ):
                    Catagory,skip = BackEnd.GUI (Write_Full_pth, identifier)
                    if (skip==1):
                        break
                    ID_values[Catagory].append(Transaction_Amount[x+1].value)
                    break


for key in ID_values:
    ID_values_summed[key] = sum(ID_values[key])

Catagories_Row = str (5)
Values_Row = str(6)

Write_workbook2 = xl.load_workbook(Write_Full_pth)
Write_sheet=Write_workbook2.get_sheet_by_name(month)
for x in range(len(ID.keys())):

    Col = letters[x+2]
    Write_sheet[Col+Catagories_Row] = list(ID.keys())[x] # write catagories
    Write_sheet[Col+Values_Row] = (ID_values_summed.get (list(ID.keys())[x])) # write catagories values


Write_workbook2.save(Write_Full_pth)


Analysis.plotting( ID_values_summed, month, Write_workbook)
Analysis.Summary(ID_values_summed,month,Write_workbook)

## TODO: make skipable as column in the sheet
## TODO: add more comments and upload to Git
